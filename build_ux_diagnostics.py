from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent

INPUT_FILE = (
    PROJECT_ROOT
    / "data"
    / "parsed"
    / "ux_metrics.xlsx"
)

OUTPUT_FILE = (
    PROJECT_ROOT
    / "reports"
    / "current"
    / "ux_report.xlsx"
)


# Outlier thresholds based on the interquartile range.
OUTLIER_IQR_MULTIPLIER = 1.5

# Friction-score weights. These add up to 1.00.
FRICTION_WEIGHTS = {
    "revisitRate": 0.35,
    "incompleteEndingRate": 0.30,
    "dropoffRate": 0.25,
    "relativePosition": 0.10,
}


def require_columns(
    dataframe: pd.DataFrame,
    sheet_name: str,
    required_columns: list[str],
) -> None:
    """
    Stop with a clear error if a required column is missing.
    """
    missing_columns = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            f"The '{sheet_name}' sheet is missing these columns: "
            f"{missing_columns}"
        )


def safe_divide(
    numerator: pd.Series,
    denominator: pd.Series,
) -> pd.Series:
    """
    Divide two Series without creating infinity values.
    """
    denominator = denominator.replace(0, np.nan)

    result = numerator / denominator

    return result.replace(
        [np.inf, -np.inf],
        np.nan,
    ).fillna(0)


def min_max_scale(
    series: pd.Series,
) -> pd.Series:
    """
    Scale a numeric Series between 0 and 1.

    If all valid values are identical, return zeros.
    """
    numeric_series = pd.to_numeric(
        series,
        errors="coerce",
    ).fillna(0)

    minimum = numeric_series.min()
    maximum = numeric_series.max()

    if maximum == minimum:
        return pd.Series(
            0.0,
            index=numeric_series.index,
        )

    return (
        numeric_series - minimum
    ) / (
        maximum - minimum
    )


def make_excel_safe(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Remove timezone metadata before writing datetimes to Excel.
    """
    safe_dataframe = dataframe.copy()

    for column_name in safe_dataframe.columns:
        column = safe_dataframe[column_name]

        if isinstance(column.dtype, pd.DatetimeTZDtype):
            safe_dataframe[column_name] = (
                column.dt.tz_localize(None)
            )

    return safe_dataframe


def calculate_iqr_bounds(
    series: pd.Series,
) -> tuple[float | None, float | None]:
    """
    Calculate lower and upper IQR outlier boundaries.
    """
    numeric_series = pd.to_numeric(
        series,
        errors="coerce",
    ).dropna()

    if numeric_series.empty:
        return None, None

    first_quartile = numeric_series.quantile(0.25)
    third_quartile = numeric_series.quantile(0.75)
    interquartile_range = (
        third_quartile - first_quartile
    )

    lower_bound = (
        first_quartile
        - OUTLIER_IQR_MULTIPLIER
        * interquartile_range
    )

    upper_bound = (
        third_quartile
        + OUTLIER_IQR_MULTIPLIER
        * interquartile_range
    )

    return lower_bound, upper_bound


def classify_friction_score(
    score: float,
) -> str:
    """
    Convert a numeric friction score into a readable category.
    """
    if score >= 70:
        return "High"

    if score >= 40:
        return "Moderate"

    return "Low"


def classify_outlier_reasons(
    row: pd.Series,
    completion_upper_bound: float | None,
    events_upper_bound: float | None,
    repeats_upper_bound: float | None,
) -> str:
    """
    Describe why a session was flagged as an outlier.
    """
    reasons = []

    completion_minutes = row.get(
        "completionMinutes"
    )

    number_of_events = row.get(
        "numberOfEvents"
    )

    repeat_event_count = row.get(
        "repeatEventCount"
    )

    if (
        completion_upper_bound is not None
        and pd.notna(completion_minutes)
        and completion_minutes
        > completion_upper_bound
    ):
        reasons.append(
            "Unusually long completion time"
        )

    if (
        events_upper_bound is not None
        and pd.notna(number_of_events)
        and number_of_events
        > events_upper_bound
    ):
        reasons.append(
            "Unusually high event count"
        )

    if (
        repeats_upper_bound is not None
        and pd.notna(repeat_event_count)
        and repeat_event_count
        > repeats_upper_bound
    ):
        reasons.append(
            "Unusually high repeat count"
        )

    return "; ".join(reasons)


def build_dropoff_analysis(
    session_summary_df: pd.DataFrame,
    question_summary_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Estimate question-level drop-off using incomplete sessions'
    final recorded question.

    Important:
    This does not prove that a question caused abandonment.
    It shows where incomplete sessions were last recorded.
    """
    question_df = question_summary_df.copy()

    incomplete_sessions_df = session_summary_df[
        session_summary_df["sessionStatus"]
        == "Incomplete"
    ].copy()

    incomplete_endings = (
        incomplete_sessions_df.groupby("lastEvent")
        .size()
        .rename("incompleteSessionsEndingHere")
        .reset_index()
        .rename(
            columns={
                "lastEvent": "questionId",
            }
        )
    )

    question_df = question_df.merge(
        incomplete_endings,
        on="questionId",
        how="left",
    )

    question_df[
        "incompleteSessionsEndingHere"
    ] = (
        question_df[
            "incompleteSessionsEndingHere"
        ]
        .fillna(0)
        .astype(int)
    )

    if "sessionsReached" not in question_df.columns:
        question_df["sessionsReached"] = 0

    question_df["dropoffRate"] = safe_divide(
        question_df[
            "incompleteSessionsEndingHere"
        ],
        question_df["sessionsReached"],
    )

    question_df["dropoffPercent"] = (
        question_df["dropoffRate"] * 100
    ).round(2)

    preferred_columns = [
        "questionId",
        "sessionsReached",
        "incompleteSessionsEndingHere",
        "dropoffPercent",
        "sessionsEndingHere",
        "averageEventOrder",
        "medianEventOrder",
    ]

    existing_columns = [
        column
        for column in preferred_columns
        if column in question_df.columns
    ]

    dropoff_df = question_df[
        existing_columns
    ].copy()

    dropoff_df = dropoff_df.sort_values(
        by=[
            "dropoffPercent",
            "incompleteSessionsEndingHere",
        ],
        ascending=[
            False,
            False,
        ],
        na_position="last",
    )

    return dropoff_df


def build_backtracking_analysis(
    question_summary_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Summarize repeated recorded visits by question.
    """
    backtracking_df = (
        question_summary_df.copy()
    )

    for column_name in [
        "sessionsReached",
        "repeatVisits",
        "totalRecordedEvents",
    ]:
        if column_name not in backtracking_df.columns:
            backtracking_df[column_name] = 0

    backtracking_df[
        "averageRepeatVisitsPerReachedSession"
    ] = safe_divide(
        backtracking_df["repeatVisits"],
        backtracking_df["sessionsReached"],
    ).round(3)

    # This is a proxy because the source data does not
    # currently identify exactly how many distinct sessions
    # contained at least one repeat.
    backtracking_df[
        "repeatIntensityPercent"
    ] = (
        backtracking_df[
            "averageRepeatVisitsPerReachedSession"
        ]
        * 100
    ).round(2)

    preferred_columns = [
        "questionId",
        "sessionsReached",
        "totalRecordedEvents",
        "repeatVisits",
        "averageRepeatVisitsPerReachedSession",
        "repeatIntensityPercent",
        "averageEventOrder",
        "sessionsEndingHere",
    ]

    existing_columns = [
        column
        for column in preferred_columns
        if column in backtracking_df.columns
    ]

    backtracking_df = backtracking_df[
        existing_columns
    ].copy()

    backtracking_df = backtracking_df.sort_values(
        by=[
            "repeatVisits",
            "repeatIntensityPercent",
        ],
        ascending=[
            False,
            False,
        ],
        na_position="last",
    )

    return backtracking_df


def build_outlier_sessions(
    session_summary_df: pd.DataFrame,
) -> tuple[pd.DataFrame, dict]:
    """
    Flag sessions with unusually high values for:
    - completion time
    - event count
    - repeat count
    """
    outlier_source_df = (
        session_summary_df.copy()
    )

    completion_lower_bound, completion_upper_bound = (
        calculate_iqr_bounds(
            outlier_source_df[
                "completionMinutes"
            ]
        )
    )

    events_lower_bound, events_upper_bound = (
        calculate_iqr_bounds(
            outlier_source_df[
                "numberOfEvents"
            ]
        )
    )

    repeats_lower_bound, repeats_upper_bound = (
        calculate_iqr_bounds(
            outlier_source_df[
                "repeatEventCount"
            ]
        )
    )

    outlier_source_df["outlierReason"] = (
        outlier_source_df.apply(
            lambda row: classify_outlier_reasons(
                row,
                completion_upper_bound,
                events_upper_bound,
                repeats_upper_bound,
            ),
            axis=1,
        )
    )

    outlier_df = outlier_source_df[
        outlier_source_df[
            "outlierReason"
        ].ne("")
    ].copy()

    preferred_columns = [
        "shortIntakeId",
        "formType",
        "sessionStatus",
        "completionMinutes",
        "numberOfEvents",
        "repeatEventCount",
        "numberOfFields",
        "firstEvent",
        "lastEvent",
        "eventPath",
        "outlierReason",
    ]

    existing_columns = [
        column
        for column in preferred_columns
        if column in outlier_df.columns
    ]

    outlier_df = outlier_df[
        existing_columns
    ].copy()

    sort_columns = [
        column
        for column in [
            "completionMinutes",
            "numberOfEvents",
            "repeatEventCount",
        ]
        if column in outlier_df.columns
    ]

    if sort_columns:
        outlier_df = outlier_df.sort_values(
            by=sort_columns,
            ascending=False,
            na_position="last",
        )

    thresholds = {
        "completionMinutesUpperBound": (
            completion_upper_bound
        ),
        "numberOfEventsUpperBound": (
            events_upper_bound
        ),
        "repeatEventCountUpperBound": (
            repeats_upper_bound
        ),
    }

    return outlier_df, thresholds


def build_form_comparison(
    session_summary_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Compare core UX metrics across form types.
    """
    form_df = (
        session_summary_df.copy()
    )

    form_df["isCompleted"] = (
        form_df["sessionStatus"]
        == "Completed"
    ).astype(int)

    form_comparison_df = (
        form_df.groupby(
            "formType",
            dropna=False,
        )
        .agg(
            totalSessions=(
                "shortIntakeId",
                "size",
            ),
            completedSessions=(
                "isCompleted",
                "sum",
            ),
            averageCompletionMinutes=(
                "completionMinutes",
                "mean",
            ),
            medianCompletionMinutes=(
                "completionMinutes",
                "median",
            ),
            averageEvents=(
                "numberOfEvents",
                "mean",
            ),
            averageUniqueFields=(
                "numberOfFields",
                "mean",
            ),
            averageRepeatEvents=(
                "repeatEventCount",
                "mean",
            ),
        )
        .reset_index()
    )

    form_comparison_df["incompleteSessions"] = (
        form_comparison_df["totalSessions"]
        - form_comparison_df[
            "completedSessions"
        ]
    )

    form_comparison_df["completionRate"] = (
        safe_divide(
            form_comparison_df[
                "completedSessions"
            ],
            form_comparison_df[
                "totalSessions"
            ],
        )
        * 100
    ).round(2)

    numeric_columns = [
        "averageCompletionMinutes",
        "medianCompletionMinutes",
        "averageEvents",
        "averageUniqueFields",
        "averageRepeatEvents",
    ]

    for column_name in numeric_columns:
        form_comparison_df[column_name] = (
            form_comparison_df[
                column_name
            ].round(2)
        )

    preferred_columns = [
        "formType",
        "totalSessions",
        "completedSessions",
        "incompleteSessions",
        "completionRate",
        "averageCompletionMinutes",
        "medianCompletionMinutes",
        "averageEvents",
        "averageUniqueFields",
        "averageRepeatEvents",
    ]

    form_comparison_df = (
        form_comparison_df[
            preferred_columns
        ]
    )

    form_comparison_df = (
        form_comparison_df.sort_values(
            by="totalSessions",
            ascending=False,
        )
    )

    return form_comparison_df


def build_friction_ranking(
    question_summary_df: pd.DataFrame,
    dropoff_df: pd.DataFrame,
    incomplete_session_count: int,
) -> pd.DataFrame:
    """
    Create a ranked question-level friction score.

    The score is a prioritization aid, not proof that a
    question is poorly designed.

    Components:
    - revisit rate
    - incomplete-session ending rate
    - drop-off rate
    - relative position in the form
    """
    friction_df = question_summary_df.copy()

    friction_df = friction_df.merge(
        dropoff_df[
            [
                "questionId",
                "incompleteSessionsEndingHere",
                "dropoffPercent",
            ]
        ],
        on="questionId",
        how="left",
    )

    fill_columns = [
        "sessionsReached",
        "repeatVisits",
        "incompleteSessionsEndingHere",
        "dropoffPercent",
        "averageEventOrder",
    ]

    for column_name in fill_columns:
        if column_name not in friction_df.columns:
            friction_df[column_name] = 0

        friction_df[column_name] = (
            pd.to_numeric(
                friction_df[column_name],
                errors="coerce",
            ).fillna(0)
        )

    friction_df["revisitRate"] = (
        safe_divide(
            friction_df["repeatVisits"],
            friction_df["sessionsReached"],
        )
        * 100
    ).round(2)

    if incomplete_session_count > 0:
        friction_df[
            "incompleteEndingRate"
        ] = (
            friction_df[
                "incompleteSessionsEndingHere"
            ]
            / incomplete_session_count
            * 100
        ).round(2)
    else:
        friction_df[
            "incompleteEndingRate"
        ] = 0.0

    friction_df[
        "revisitComponent"
    ] = min_max_scale(
        friction_df["revisitRate"]
    )

    friction_df[
        "incompleteEndingComponent"
    ] = min_max_scale(
        friction_df[
            "incompleteEndingRate"
        ]
    )

    friction_df[
        "dropoffComponent"
    ] = min_max_scale(
        friction_df["dropoffPercent"]
    )

    friction_df[
        "relativePositionComponent"
    ] = min_max_scale(
        friction_df[
            "averageEventOrder"
        ]
    )

    friction_df["frictionScore"] = (
        (
            friction_df[
                "revisitComponent"
            ]
            * FRICTION_WEIGHTS[
                "revisitRate"
            ]
        )
        + (
            friction_df[
                "incompleteEndingComponent"
            ]
            * FRICTION_WEIGHTS[
                "incompleteEndingRate"
            ]
        )
        + (
            friction_df[
                "dropoffComponent"
            ]
            * FRICTION_WEIGHTS[
                "dropoffRate"
            ]
        )
        + (
            friction_df[
                "relativePositionComponent"
            ]
            * FRICTION_WEIGHTS[
                "relativePosition"
            ]
        )
    ) * 100

    friction_df["frictionScore"] = (
        friction_df[
            "frictionScore"
        ].round(2)
    )

    friction_df["frictionLevel"] = (
        friction_df[
            "frictionScore"
        ].apply(
            classify_friction_score
        )
    )

    friction_df["primarySignal"] = (
        friction_df.apply(
            identify_primary_signal,
            axis=1,
        )
    )

    preferred_columns = [
        "questionId",
        "frictionScore",
        "frictionLevel",
        "primarySignal",
        "sessionsReached",
        "repeatVisits",
        "revisitRate",
        "incompleteSessionsEndingHere",
        "incompleteEndingRate",
        "dropoffPercent",
        "averageEventOrder",
        "medianEventOrder",
        "sessionsEndingHere",
    ]

    existing_columns = [
        column
        for column in preferred_columns
        if column in friction_df.columns
    ]

    friction_df = friction_df[
        existing_columns
    ].copy()

    friction_df = friction_df.sort_values(
        by=[
            "frictionScore",
            "sessionsReached",
        ],
        ascending=[
            False,
            False,
        ],
        na_position="last",
    )

    friction_df.insert(
        0,
        "frictionRank",
        range(
            1,
            len(friction_df) + 1,
        ),
    )

    return friction_df


def identify_primary_signal(
    row: pd.Series,
) -> str:
    """
    Identify which friction component contributes most
    strongly for a question.
    """
    components = {
        "Repeated recorded visits": (
            row.get(
                "revisitComponent",
                0,
            )
            * FRICTION_WEIGHTS[
                "revisitRate"
            ]
        ),
        "Incomplete sessions ending here": (
            row.get(
                "incompleteEndingComponent",
                0,
            )
            * FRICTION_WEIGHTS[
                "incompleteEndingRate"
            ]
        ),
        "Question-level drop-off": (
            row.get(
                "dropoffComponent",
                0,
            )
            * FRICTION_WEIGHTS[
                "dropoffRate"
            ]
        ),
        "Late position in flow": (
            row.get(
                "relativePositionComponent",
                0,
            )
            * FRICTION_WEIGHTS[
                "relativePosition"
            ]
        ),
    }

    highest_signal = max(
        components,
        key=components.get,
    )

    if components[highest_signal] == 0:
        return "No elevated signal"

    return highest_signal


def build_methodology_sheet(
    thresholds: dict,
) -> pd.DataFrame:
    """
    Document how each diagnostic metric is calculated.
    """
    return pd.DataFrame(
        [
            {
                "metric": "Drop-off percentage",
                "calculation": (
                    "Incomplete sessions whose final recorded "
                    "event was the question, divided by sessions "
                    "that reached the question."
                ),
                "interpretation": (
                    "A directional indicator of where incomplete "
                    "sessions ended. It does not prove the question "
                    "caused abandonment."
                ),
            },
            {
                "metric": "Repeat visits",
                "calculation": (
                    "Recorded appearances of a question after its "
                    "first appearance within each session."
                ),
                "interpretation": (
                    "May indicate backtracking, correction, review, "
                    "or expected navigation."
                ),
            },
            {
                "metric": "Friction score",
                "calculation": (
                    "35% normalized revisit rate + "
                    "30% normalized incomplete-ending rate + "
                    "25% normalized drop-off rate + "
                    "10% normalized relative form position."
                ),
                "interpretation": (
                    "A relative prioritization score from 0 to 100. "
                    "It should be validated through session review "
                    "and user research."
                ),
            },
            {
                "metric": "Completion-time outlier",
                "calculation": (
                    f"Above the IQR upper boundary of "
                    f"{format_optional_number(thresholds.get('completionMinutesUpperBound'))} "
                    f"minutes."
                ),
                "interpretation": (
                    "Identifies unusually long sessions. Long time "
                    "may reflect interruption rather than UX friction."
                ),
            },
            {
                "metric": "Event-count outlier",
                "calculation": (
                    f"Above the IQR upper boundary of "
                    f"{format_optional_number(thresholds.get('numberOfEventsUpperBound'))} "
                    f"events."
                ),
                "interpretation": (
                    "Identifies sessions with unusually high "
                    "interaction volume."
                ),
            },
            {
                "metric": "Repeat-count outlier",
                "calculation": (
                    f"Above the IQR upper boundary of "
                    f"{format_optional_number(thresholds.get('repeatEventCountUpperBound'))} "
                    f"repeat events."
                ),
                "interpretation": (
                    "Identifies sessions with unusually high "
                    "repeated interactions."
                ),
            },
        ]
    )


def format_optional_number(
    value: float | None,
) -> str:
    """
    Format an optional numeric value for readable text.
    """
    if value is None or pd.isna(value):
        return "not available"

    return f"{value:.2f}"


def build_executive_summary(
    session_summary_df: pd.DataFrame,
    friction_df: pd.DataFrame,
    dropoff_df: pd.DataFrame,
    outlier_df: pd.DataFrame,
    form_comparison_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Produce plain-English findings for leadership review.
    """
    summary_rows = []

    total_sessions = len(
        session_summary_df
    )

    completed_sessions = int(
        (
            session_summary_df[
                "sessionStatus"
            ]
            == "Completed"
        ).sum()
    )

    incomplete_sessions = (
        total_sessions
        - completed_sessions
    )

    completion_rate = (
        completed_sessions
        / total_sessions
        * 100
        if total_sessions
        else 0
    )

    valid_completion_times = (
        pd.to_numeric(
            session_summary_df[
                "completionMinutes"
            ],
            errors="coerce",
        ).dropna()
    )

    average_completion = (
        valid_completion_times.mean()
        if not valid_completion_times.empty
        else np.nan
    )

    median_completion = (
        valid_completion_times.median()
        if not valid_completion_times.empty
        else np.nan
    )

    summary_rows.append(
        {
            "category": "Volume",
            "finding": (
                f"{total_sessions} sessions were analyzed. "
                f"{completed_sessions} were completed and "
                f"{incomplete_sessions} were incomplete."
            ),
            "recommendedAction": (
                "Track this baseline over time and segment it "
                "by deployment, date range, and form type."
            ),
        }
    )

    summary_rows.append(
        {
            "category": "Completion",
            "finding": (
                f"The overall completion rate was "
                f"{completion_rate:.1f}%."
            ),
            "recommendedAction": (
                "Investigate form types or questions associated "
                "with lower completion."
            ),
        }
    )

    if pd.notna(average_completion):
        summary_rows.append(
            {
                "category": "Timing",
                "finding": (
                    f"Average completion time was "
                    f"{average_completion:.2f} minutes, "
                    f"with a median of "
                    f"{median_completion:.2f} minutes."
                ),
                "recommendedAction": (
                    "Use the median as the stronger baseline when "
                    "long interrupted sessions skew the average."
                ),
            }
        )

    if not friction_df.empty:
        top_friction = friction_df.iloc[0]

        summary_rows.append(
            {
                "category": "Highest-priority question",
                "finding": (
                    f"'{top_friction['questionId']}' had the "
                    f"highest relative friction score at "
                    f"{top_friction['frictionScore']:.1f}/100. "
                    f"Its strongest signal was "
                    f"{str(top_friction['primarySignal']).lower()}."
                ),
                "recommendedAction": (
                    "Review the wording, response options, validation "
                    "rules, and surrounding navigation for this "
                    "question. Validate with session review before "
                    "making a product change."
                ),
            }
        )

    high_friction_count = int(
        (
            friction_df[
                "frictionLevel"
            ]
            == "High"
        ).sum()
    ) if not friction_df.empty else 0

    summary_rows.append(
        {
            "category": "Friction distribution",
            "finding": (
                f"{high_friction_count} questions were classified "
                f"as high-friction relative to the other questions "
                f"in this dataset."
            ),
            "recommendedAction": (
                "Prioritize the highest-ranked questions for "
                "qualitative review and usability testing."
            ),
        }
    )

    valid_dropoff_df = dropoff_df[
        dropoff_df[
            "incompleteSessionsEndingHere"
        ] > 0
    ]

    if not valid_dropoff_df.empty:
        top_dropoff = (
            valid_dropoff_df.iloc[0]
        )

        summary_rows.append(
            {
                "category": "Incomplete-session endpoint",
                "finding": (
                    f"'{top_dropoff['questionId']}' was the final "
                    f"recorded question for "
                    f"{int(top_dropoff['incompleteSessionsEndingHere'])} "
                    f"incomplete sessions, representing a "
                    f"{top_dropoff['dropoffPercent']:.1f}% "
                    f"question-level drop-off rate."
                ),
                "recommendedAction": (
                    "Inspect incomplete sessions ending at this "
                    "question and confirm whether the endpoint reflects "
                    "confusion, expected stopping, or instrumentation."
                ),
            }
        )

    summary_rows.append(
        {
            "category": "Outliers",
            "finding": (
                f"{len(outlier_df)} sessions were flagged as "
                f"unusually long, event-heavy, or repetitive."
            ),
            "recommendedAction": (
                "Manually review these sessions before treating "
                "them as representative UX behavior."
            ),
        }
    )

    if len(form_comparison_df) >= 2:
        valid_form_times = (
            form_comparison_df.dropna(
                subset=[
                    "averageCompletionMinutes"
                ]
            )
        )

        if len(valid_form_times) >= 2:
            slowest_form = (
                valid_form_times.sort_values(
                    by="averageCompletionMinutes",
                    ascending=False,
                ).iloc[0]
            )

            fastest_form = (
                valid_form_times.sort_values(
                    by="averageCompletionMinutes",
                    ascending=True,
                ).iloc[0]
            )

            fastest_time = (
                fastest_form[
                    "averageCompletionMinutes"
                ]
            )

            slowest_time = (
                slowest_form[
                    "averageCompletionMinutes"
                ]
            )

            if (
                pd.notna(fastest_time)
                and fastest_time > 0
            ):
                percent_longer = (
                    (
                        slowest_time
                        - fastest_time
                    )
                    / fastest_time
                    * 100
                )

                summary_rows.append(
                    {
                        "category": "Form comparison",
                        "finding": (
                            f"'{slowest_form['formType']}' took "
                            f"{percent_longer:.1f}% longer on average "
                            f"than '{fastest_form['formType']}'."
                        ),
                        "recommendedAction": (
                            "Compare question count, question design, "
                            "and user population before attributing "
                            "the difference to UX quality."
                        ),
                    }
                )

    summary_rows.append(
        {
            "category": "Interpretation note",
            "finding": (
                "These diagnostics identify behavioral patterns, "
                "not confirmed causes."
            ),
            "recommendedAction": (
                "Pair the quantitative findings with session review, "
                "user observation, and qualitative feedback."
            ),
        }
    )

    return pd.DataFrame(
        summary_rows
    )


def auto_fit_and_format(
    writer: pd.ExcelWriter,
) -> None:
    """
    Apply readable formatting to all workbook sheets.
    """
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            maximum_length = 0
            column_number = (
                column_cells[0].column
            )

            column_letter = (
                get_column_letter(
                    column_number
                )
            )

            for cell in column_cells:
                if cell.value is None:
                    continue

                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                60,
            )

        if worksheet.title == "Friction Ranking":
            header_lookup = {
                cell.value: cell.column
                for cell in worksheet[1]
            }

            friction_score_column = (
                header_lookup.get(
                    "frictionScore"
                )
            )

            if friction_score_column:
                column_letter = (
                    get_column_letter(
                        friction_score_column
                    )
                )

                worksheet.conditional_formatting.add(
                    (
                        f"{column_letter}2:"
                        f"{column_letter}"
                        f"{worksheet.max_row}"
                    ),
                    ColorScaleRule(
                        start_type="min",
                        start_color="63BE7B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="F8696B",
                    ),
                )

        if worksheet.title in {
            "Drop-off Analysis",
            "Backtracking Analysis",
        }:
            header_lookup = {
                cell.value: cell.column
                for cell in worksheet[1]
            }

            for metric_name in [
                "dropoffPercent",
                "repeatIntensityPercent",
            ]:
                metric_column = (
                    header_lookup.get(
                        metric_name
                    )
                )

                if metric_column:
                    column_letter = (
                        get_column_letter(
                            metric_column
                        )
                    )

                    worksheet.conditional_formatting.add(
                        (
                            f"{column_letter}2:"
                            f"{column_letter}"
                            f"{worksheet.max_row}"
                        ),
                        ColorScaleRule(
                            start_type="min",
                            start_color="63BE7B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="F8696B",
                        ),
                    )


def main() -> None:
    print("Reading UX metrics workbook...")

    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Could not find: "
            f"{INPUT_FILE.resolve()}"
        )

    session_summary_df = pd.read_excel(
        INPUT_FILE,
        sheet_name="Session Summary",
    )

    question_summary_df = pd.read_excel(
        INPUT_FILE,
        sheet_name="Question Summary",
    )

    session_paths_df = pd.read_excel(
        INPUT_FILE,
        sheet_name="Session Paths",
    )

    summary_metrics_df = pd.read_excel(
        INPUT_FILE,
        sheet_name="Summary Metrics",
    )

    require_columns(
        session_summary_df,
        "Session Summary",
        [
            "shortIntakeId",
            "formType",
            "sessionStatus",
            "completionMinutes",
            "numberOfEvents",
            "repeatEventCount",
            "numberOfFields",
            "lastEvent",
        ],
    )

    require_columns(
        question_summary_df,
        "Question Summary",
        [
            "questionId",
            "sessionsReached",
            "repeatVisits",
            "averageEventOrder",
        ],
    )

    incomplete_session_count = int(
        (
            session_summary_df[
                "sessionStatus"
            ]
            == "Incomplete"
        ).sum()
    )

    print("Building drop-off analysis...")

    dropoff_df = build_dropoff_analysis(
        session_summary_df,
        question_summary_df,
    )

    print("Building backtracking analysis...")

    backtracking_df = (
        build_backtracking_analysis(
            question_summary_df
        )
    )

    print("Finding outlier sessions...")

    outlier_df, thresholds = (
        build_outlier_sessions(
            session_summary_df
        )
    )

    print("Comparing form types...")

    form_comparison_df = (
        build_form_comparison(
            session_summary_df
        )
    )

    print("Calculating friction scores...")

    friction_df = (
        build_friction_ranking(
            question_summary_df,
            dropoff_df,
            incomplete_session_count,
        )
    )

    print("Writing executive summary...")

    executive_summary_df = (
        build_executive_summary(
            session_summary_df,
            friction_df,
            dropoff_df,
            outlier_df,
            form_comparison_df,
        )
    )

    methodology_df = (
        build_methodology_sheet(
            thresholds
        )
    )

    output_dataframes = {
        "Executive Summary": (
            executive_summary_df
        ),
        "Friction Ranking": (
            friction_df
        ),
        "Drop-off Analysis": (
            dropoff_df
        ),
        "Backtracking Analysis": (
            backtracking_df
        ),
        "Outlier Sessions": (
            outlier_df
        ),
        "Form Comparison": (
            form_comparison_df
        ),
        "Summary Metrics": (
            summary_metrics_df
        ),
        "Session Summary": (
            session_summary_df
        ),
        "Question Summary": (
            question_summary_df
        ),
        "Session Paths": (
            session_paths_df
        ),
        "Methodology": (
            methodology_df
        ),
    }

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    print("Writing diagnostics workbook...")

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        for (
            sheet_name,
            dataframe,
        ) in output_dataframes.items():
            safe_dataframe = (
                make_excel_safe(
                    dataframe
                )
            )

            safe_dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        auto_fit_and_format(
            writer
        )

    print()
    print(
        "UX diagnostics workbook complete."
    )

    print(
        f"Sessions analyzed: "
        f"{len(session_summary_df)}"
    )

    print(
        f"Questions ranked: "
        f"{len(friction_df)}"
    )

    print(
        f"Outlier sessions: "
        f"{len(outlier_df)}"
    )

    print(
        f"Saved to: "
        f"{OUTPUT_FILE.resolve()}"
    )


if __name__ == "__main__":
    main()