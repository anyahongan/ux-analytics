from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent

PARSED_FILE = (
    PROJECT_ROOT
    / "data"
    / "parsed"
    / "master_parsed.xlsx"
)

METRICS_FILE = (
    PROJECT_ROOT
    / "data"
    / "parsed"
    / "ux_metrics.xlsx"
)

DIAGNOSTICS_FILE = (
    PROJECT_ROOT
    / "reports"
    / "current"
    / "ux_report.xlsx"
)

OUTPUT_FILE = (
    PROJECT_ROOT
    / "reports"
    / "current"
    / "ux_validation_report.xlsx"
)


# Questions reached by fewer than this many sessions
# will receive a low-sample warning.
LOW_SAMPLE_THRESHOLD = 10


def require_file(
    file_path: Path,
) -> None:
    """
    Stop with a clear message when a required workbook is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"Could not find required file: "
            f"{file_path.resolve()}"
        )


def require_columns(
    dataframe: pd.DataFrame,
    sheet_name: str,
    required_columns: list[str],
) -> None:
    """
    Stop when an expected column is missing.
    """
    missing_columns = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            f"The '{sheet_name}' sheet is missing: "
            f"{missing_columns}"
        )


def prepare_datetime_column(
    dataframe: pd.DataFrame,
    column_name: str,
) -> None:
    """
    Convert a column to timezone-naive UTC datetime values.

    This allows timestamps from different sources to be
    compared consistently.
    """
    if column_name not in dataframe.columns:
        return

    dataframe[column_name] = pd.to_datetime(
        dataframe[column_name],
        errors="coerce",
        utc=True,
    ).dt.tz_localize(None)


def make_excel_safe(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Remove timezone metadata before writing to Excel.
    """
    safe_dataframe = dataframe.copy()

    for column_name in safe_dataframe.columns:
        column = safe_dataframe[column_name]

        if isinstance(
            column.dtype,
            pd.DatetimeTZDtype,
        ):
            safe_dataframe[column_name] = (
                column.dt.tz_localize(None)
            )

    return safe_dataframe


def add_issue(
    issue_rows: list[dict],
    issue_type: str,
    severity: str,
    record_id,
    description: str,
    recommended_action: str,
) -> None:
    """
    Add one standardized issue row.
    """
    issue_rows.append(
        {
            "issueType": issue_type,
            "severity": severity,
            "recordId": record_id,
            "description": description,
            "recommendedAction": recommended_action,
        }
    )


def build_session_issues(
    sessions_df: pd.DataFrame,
    session_summary_df: pd.DataFrame,
    events_df: pd.DataFrame,
    fields_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Check session-level consistency and completeness.
    """
    issues = []

    session_ids = set(
        sessions_df["shortIntakeId"]
        .dropna()
        .astype(str)
    )

    # Duplicate session identifiers
    duplicate_mask = sessions_df[
        "shortIntakeId"
    ].duplicated(
        keep=False
    )

    duplicate_sessions = sessions_df[
        duplicate_mask
    ]

    for _, row in duplicate_sessions.iterrows():
        add_issue(
            issues,
            issue_type="Duplicate session ID",
            severity="High",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "The same shortIntakeId appears more than "
                "once in the Sessions sheet."
            ),
            recommended_action=(
                "Confirm whether these are duplicate exports, "
                "multiple records from one intake, or expected "
                "source-system behavior."
            ),
        )

    # Missing IDs
    missing_id_rows = sessions_df[
        sessions_df["shortIntakeId"].isna()
        | sessions_df[
            "shortIntakeId"
        ].astype(str).str.strip().eq("")
    ]

    for index, _ in missing_id_rows.iterrows():
        add_issue(
            issues,
            issue_type="Missing session ID",
            severity="High",
            record_id=f"Sessions row {index + 2}",
            description=(
                "The session does not have a usable "
                "shortIntakeId."
            ),
            recommended_action=(
                "Inspect the original intakeId and RowKey "
                "in the raw export."
            ),
        )

    # Missing timestamps
    missing_start_rows = sessions_df[
        sessions_df["startedAt"].isna()
    ]

    for _, row in missing_start_rows.iterrows():
        add_issue(
            issues,
            issue_type="Missing start time",
            severity="High",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "The session has no valid startedAt timestamp."
            ),
            recommended_action=(
                "Check whether the source record is incomplete "
                "or the timestamp failed during parsing."
            ),
        )

    # Submitted before started
    invalid_time_rows = sessions_df[
        sessions_df["startedAt"].notna()
        & sessions_df["submittedAt"].notna()
        & (
            sessions_df["submittedAt"]
            < sessions_df["startedAt"]
        )
    ]

    for _, row in invalid_time_rows.iterrows():
        add_issue(
            issues,
            issue_type="Invalid timestamp order",
            severity="High",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "submittedAt occurs before startedAt."
            ),
            recommended_action=(
                "Inspect timezone handling and source timestamps."
            ),
        )

    # Negative completion values
    if "completionSeconds" in sessions_df.columns:
        negative_completion_rows = sessions_df[
            pd.to_numeric(
                sessions_df[
                    "completionSeconds"
                ],
                errors="coerce",
            ) < 0
        ]

        for _, row in negative_completion_rows.iterrows():
            add_issue(
                issues,
                issue_type="Negative completion time",
                severity="High",
                record_id=row.get(
                    "shortIntakeId"
                ),
                description=(
                    "The calculated completion time is negative."
                ),
                recommended_action=(
                    "Review timestamp parsing and timezone logic."
                ),
            )

    # Completion time mismatch
    calculated_seconds = (
        sessions_df["submittedAt"]
        - sessions_df["startedAt"]
    ).dt.total_seconds()

    stored_seconds = pd.to_numeric(
        sessions_df.get(
            "completionSeconds",
            pd.Series(
                np.nan,
                index=sessions_df.index,
            ),
        ),
        errors="coerce",
    )

    mismatch_mask = (
        calculated_seconds.notna()
        & stored_seconds.notna()
        & (
            calculated_seconds
            - stored_seconds
        ).abs().gt(1)
    )

    mismatch_rows = sessions_df[
        mismatch_mask
    ]

    for index, row in mismatch_rows.iterrows():
        add_issue(
            issues,
            issue_type="Completion calculation mismatch",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "The stored completionSeconds differs from "
                "submittedAt minus startedAt by more than "
                "one second."
            ),
            recommended_action=(
                "Recalculate completion time during parsing "
                "rather than relying on a stored value."
            ),
        )

    # Sessions with no matching events
    event_session_ids = set(
        events_df["shortIntakeId"]
        .dropna()
        .astype(str)
    )

    sessions_without_events = (
        session_ids - event_session_ids
    )

    for session_id in sorted(
        sessions_without_events
    ):
        add_issue(
            issues,
            issue_type="Session has no events",
            severity="Medium",
            record_id=session_id,
            description=(
                "The session exists, but no matching Events "
                "rows were found."
            ),
            recommended_action=(
                "Confirm whether zero-event sessions are expected "
                "or indicate missing instrumentation."
            ),
        )

    # Sessions with no matching field records
    field_session_ids = set(
        fields_df["shortIntakeId"]
        .dropna()
        .astype(str)
    )

    sessions_without_fields = (
        session_ids - field_session_ids
    )

    for session_id in sorted(
        sessions_without_fields
    ):
        add_issue(
            issues,
            issue_type="Session has no field records",
            severity="Low",
            record_id=session_id,
            description=(
                "The session exists, but no matching Fields "
                "rows were found."
            ),
            recommended_action=(
                "Confirm whether the user made no field changes "
                "or whether fieldsJson was missing."
            ),
        )

    # Session Summary status checks
    if {
        "sessionStatus",
        "submittedAt",
    }.issubset(
        session_summary_df.columns
    ):
        completed_without_submission = (
            session_summary_df[
                (
                    session_summary_df[
                        "sessionStatus"
                    ]
                    == "Completed"
                )
                & session_summary_df[
                    "submittedAt"
                ].isna()
            ]
        )

        for _, row in (
            completed_without_submission.iterrows()
        ):
            add_issue(
                issues,
                issue_type="Status mismatch",
                severity="High",
                record_id=row.get(
                    "shortIntakeId"
                ),
                description=(
                    "The session is labeled Completed but "
                    "has no submittedAt timestamp."
                ),
                recommended_action=(
                    "Correct the session-status calculation."
                ),
            )

        incomplete_with_submission = (
            session_summary_df[
                (
                    session_summary_df[
                        "sessionStatus"
                    ]
                    == "Incomplete"
                )
                & session_summary_df[
                    "submittedAt"
                ].notna()
            ]
        )

        for _, row in (
            incomplete_with_submission.iterrows()
        ):
            add_issue(
                issues,
                issue_type="Status mismatch",
                severity="High",
                record_id=row.get(
                    "shortIntakeId"
                ),
                description=(
                    "The session is labeled Incomplete but "
                    "has a submittedAt timestamp."
                ),
                recommended_action=(
                    "Correct the session-status calculation."
                ),
            )

    return pd.DataFrame(
        issues,
        columns=[
            "issueType",
            "severity",
            "recordId",
            "description",
            "recommendedAction",
        ],
    )


def build_event_issues(
    sessions_df: pd.DataFrame,
    events_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Check event-level identifiers, order, and session matching.
    """
    issues = []

    session_ids = set(
        sessions_df["shortIntakeId"]
        .dropna()
        .astype(str)
    )

    # Events with no matching session
    for _, row in events_df.iterrows():
        event_session_id = row.get(
            "shortIntakeId"
        )

        if (
            pd.notna(event_session_id)
            and str(event_session_id)
            not in session_ids
        ):
            add_issue(
                issues,
                issue_type="Orphan event",
                severity="High",
                record_id=event_session_id,
                description=(
                    "An event references a session that does "
                    "not exist in the Sessions sheet."
                ),
                recommended_action=(
                    "Inspect ID-shortening logic and source exports."
                ),
            )

    # Missing question IDs
    blank_link_mask = (
        events_df["linkId"].isna()
        | events_df["linkId"]
        .astype(str)
        .str.strip()
        .eq("")
    )

    blank_link_rows = events_df[
        blank_link_mask
    ]

    for index, row in blank_link_rows.iterrows():
        add_issue(
            issues,
            issue_type="Missing event question ID",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId",
                f"Events row {index + 2}",
            ),
            description=(
                "The event has no usable linkId."
            ),
            recommended_action=(
                "Inspect eventsJson in the raw export."
            ),
        )

    # Missing event order
    numeric_event_order = pd.to_numeric(
        events_df["eventOrder"],
        errors="coerce",
    )

    missing_order_rows = events_df[
        numeric_event_order.isna()
    ]

    for index, row in (
        missing_order_rows.iterrows()
    ):
        add_issue(
            issues,
            issue_type="Missing event order",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId",
                f"Events row {index + 2}",
            ),
            description=(
                "The event does not have a valid eventOrder."
            ),
            recommended_action=(
                "Confirm the parser assigns event order based "
                "on the original JSON sequence."
            ),
        )

    # Duplicate event order within session
    order_check_df = events_df.copy()

    order_check_df["numericEventOrder"] = (
        numeric_event_order
    )

    duplicate_order_mask = (
        order_check_df.duplicated(
            subset=[
                "shortIntakeId",
                "numericEventOrder",
            ],
            keep=False,
        )
        & order_check_df[
            "numericEventOrder"
        ].notna()
    )

    duplicate_order_rows = (
        order_check_df[
            duplicate_order_mask
        ]
    )

    for _, row in (
        duplicate_order_rows.iterrows()
    ):
        add_issue(
            issues,
            issue_type="Duplicate event order",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                f"Event order "
                f"{row.get('numericEventOrder')} "
                f"appears more than once in the session."
            ),
            recommended_action=(
                "Inspect the source JSON and event-order "
                "assignment logic."
            ),
        )

    # Event timestamps outside session boundaries
    session_times_df = sessions_df[
        [
            "shortIntakeId",
            "startedAt",
            "submittedAt",
        ]
    ].copy()

    timestamp_check_df = events_df.merge(
        session_times_df,
        on="shortIntakeId",
        how="left",
    )

    before_start_rows = timestamp_check_df[
        timestamp_check_df[
            "eventTime"
        ].notna()
        & timestamp_check_df[
            "startedAt"
        ].notna()
        & (
            timestamp_check_df[
                "eventTime"
            ]
            < timestamp_check_df[
                "startedAt"
            ]
        )
    ]

    for _, row in before_start_rows.iterrows():
        add_issue(
            issues,
            issue_type="Event before session start",
            severity="High",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "An event timestamp occurs before startedAt."
            ),
            recommended_action=(
                "Review source timestamps and timezone handling."
            ),
        )

    after_submission_rows = timestamp_check_df[
        timestamp_check_df[
            "eventTime"
        ].notna()
        & timestamp_check_df[
            "submittedAt"
        ].notna()
        & (
            timestamp_check_df[
                "eventTime"
            ]
            > timestamp_check_df[
                "submittedAt"
            ]
        )
    ]

    for _, row in (
        after_submission_rows.iterrows()
    ):
        add_issue(
            issues,
            issue_type="Event after submission",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                "An event timestamp occurs after submittedAt."
            ),
            recommended_action=(
                "Determine whether post-submission events are "
                "expected or indicate timing inconsistency."
            ),
        )

    return pd.DataFrame(
        issues,
        columns=[
            "issueType",
            "severity",
            "recordId",
            "description",
            "recommendedAction",
        ],
    )


def build_field_issues(
    sessions_df: pd.DataFrame,
    fields_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Check field-level identifiers and session matching.
    """
    issues = []

    session_ids = set(
        sessions_df["shortIntakeId"]
        .dropna()
        .astype(str)
    )

    # Orphan field records
    for _, row in fields_df.iterrows():
        field_session_id = row.get(
            "shortIntakeId"
        )

        if (
            pd.notna(field_session_id)
            and str(field_session_id)
            not in session_ids
        ):
            add_issue(
                issues,
                issue_type="Orphan field record",
                severity="High",
                record_id=field_session_id,
                description=(
                    "A field record references a session that "
                    "does not exist in the Sessions sheet."
                ),
                recommended_action=(
                    "Inspect intake ID transformations and "
                    "source exports."
                ),
            )

    # Missing field IDs
    blank_field_mask = (
        fields_df["fieldId"].isna()
        | fields_df["fieldId"]
        .astype(str)
        .str.strip()
        .eq("")
    )

    blank_field_rows = fields_df[
        blank_field_mask
    ]

    for index, row in blank_field_rows.iterrows():
        add_issue(
            issues,
            issue_type="Missing field ID",
            severity="Medium",
            record_id=row.get(
                "shortIntakeId",
                f"Fields row {index + 2}",
            ),
            description=(
                "The field record has no usable fieldId."
            ),
            recommended_action=(
                "Inspect fieldsJson in the raw export."
            ),
        )

    # Duplicate field record per session
    duplicate_field_mask = (
        fields_df.duplicated(
            subset=[
                "shortIntakeId",
                "fieldId",
            ],
            keep=False,
        )
        & fields_df[
            "fieldId"
        ].notna()
    )

    duplicate_field_rows = fields_df[
        duplicate_field_mask
    ]

    for _, row in (
        duplicate_field_rows.iterrows()
    ):
        add_issue(
            issues,
            issue_type="Duplicate field record",
            severity="Low",
            record_id=row.get(
                "shortIntakeId"
            ),
            description=(
                f"The field '{row.get('fieldId')}' appears "
                f"more than once for the same session."
            ),
            recommended_action=(
                "Confirm whether duplicate field entries are "
                "expected or should be consolidated."
            ),
        )

    return pd.DataFrame(
        issues,
        columns=[
            "issueType",
            "severity",
            "recordId",
            "description",
            "recommendedAction",
        ],
    )


def build_question_warnings(
    question_summary_df: pd.DataFrame,
    friction_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Warn when question-level metrics rely on small samples.
    """
    warnings = []

    question_summary_df = (
        question_summary_df.copy()
    )

    question_summary_df[
        "sessionsReached"
    ] = pd.to_numeric(
        question_summary_df[
            "sessionsReached"
        ],
        errors="coerce",
    ).fillna(0)

    low_sample_questions = (
        question_summary_df[
            question_summary_df[
                "sessionsReached"
            ] < LOW_SAMPLE_THRESHOLD
        ]
    )

    for _, row in (
        low_sample_questions.iterrows()
    ):
        warnings.append(
            {
                "questionId": row.get(
                    "questionId"
                ),
                "warningType": (
                    "Low sample size"
                ),
                "severity": "Medium",
                "sessionsReached": int(
                    row.get(
                        "sessionsReached",
                        0,
                    )
                ),
                "frictionScore": None,
                "description": (
                    f"Fewer than "
                    f"{LOW_SAMPLE_THRESHOLD} sessions "
                    f"reached this question."
                ),
                "recommendedAction": (
                    "Treat percentages and rankings as "
                    "directional until more sessions are collected."
                ),
            }
        )

    if not friction_df.empty:
        friction_check_df = (
            friction_df.copy()
        )

        friction_check_df[
            "sessionsReached"
        ] = pd.to_numeric(
            friction_check_df[
                "sessionsReached"
            ],
            errors="coerce",
        ).fillna(0)

        friction_check_df[
            "frictionScore"
        ] = pd.to_numeric(
            friction_check_df[
                "frictionScore"
            ],
            errors="coerce",
        ).fillna(0)

        high_score_low_sample = (
            friction_check_df[
                (
                    friction_check_df[
                        "frictionScore"
                    ] >= 70
                )
                & (
                    friction_check_df[
                        "sessionsReached"
                    ]
                    < LOW_SAMPLE_THRESHOLD
                )
            ]
        )

        for _, row in (
            high_score_low_sample.iterrows()
        ):
            warnings.append(
                {
                    "questionId": row.get(
                        "questionId"
                    ),
                    "warningType": (
                        "High friction score with low sample"
                    ),
                    "severity": "High",
                    "sessionsReached": int(
                        row.get(
                            "sessionsReached",
                            0,
                        )
                    ),
                    "frictionScore": row.get(
                        "frictionScore"
                    ),
                    "description": (
                        "The question ranks as high-friction, "
                        "but relatively few sessions reached it."
                    ),
                    "recommendedAction": (
                        "Do not prioritize a redesign based on "
                        "this score alone. Review sessions and "
                        "collect additional data."
                    ),
                }
            )

    return pd.DataFrame(
        warnings,
        columns=[
            "questionId",
            "warningType",
            "severity",
            "sessionsReached",
            "frictionScore",
            "description",
            "recommendedAction",
        ],
    )


def build_validation_summary(
    session_issues_df: pd.DataFrame,
    event_issues_df: pd.DataFrame,
    field_issues_df: pd.DataFrame,
    question_warnings_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build a compact summary of validation results.
    """
    all_issues_df = pd.concat(
        [
            session_issues_df.assign(
                area="Sessions"
            ),
            event_issues_df.assign(
                area="Events"
            ),
            field_issues_df.assign(
                area="Fields"
            ),
            question_warnings_df.rename(
                columns={
                    "questionId": "recordId",
                    "warningType": "issueType",
                }
            ).assign(
                area="Questions"
            ),
        ],
        ignore_index=True,
        sort=False,
    )

    total_issues = len(
        all_issues_df
    )

    high_issues = int(
        (
            all_issues_df["severity"]
            == "High"
        ).sum()
    ) if not all_issues_df.empty else 0

    medium_issues = int(
        (
            all_issues_df["severity"]
            == "Medium"
        ).sum()
    ) if not all_issues_df.empty else 0

    low_issues = int(
        (
            all_issues_df["severity"]
            == "Low"
        ).sum()
    ) if not all_issues_df.empty else 0

    if high_issues > 0:
        overall_status = (
            "Review required"
        )
    elif medium_issues > 0:
        overall_status = (
            "Usable with cautions"
        )
    else:
        overall_status = (
            "Validation passed"
        )

    summary_rows = [
        {
            "metric": "Overall status",
            "value": overall_status,
            "interpretation": (
                "High-severity issues should be investigated "
                "before using the affected metrics."
            ),
        },
        {
            "metric": "Total validation findings",
            "value": total_issues,
            "interpretation": (
                "Includes data errors and analytical warnings."
            ),
        },
        {
            "metric": "High-severity findings",
            "value": high_issues,
            "interpretation": (
                "Potential data-integrity or interpretation risks."
            ),
        },
        {
            "metric": "Medium-severity findings",
            "value": medium_issues,
            "interpretation": (
                "Items that should be reviewed but may not "
                "invalidate the full dataset."
            ),
        },
        {
            "metric": "Low-severity findings",
            "value": low_issues,
            "interpretation": (
                "Minor anomalies or expected edge cases."
            ),
        },
        {
            "metric": "Session findings",
            "value": len(
                session_issues_df
            ),
            "interpretation": (
                "Issues related to session identifiers, "
                "timestamps, status, or completeness."
            ),
        },
        {
            "metric": "Event findings",
            "value": len(
                event_issues_df
            ),
            "interpretation": (
                "Issues related to event order, timestamps, "
                "question IDs, or session matching."
            ),
        },
        {
            "metric": "Field findings",
            "value": len(
                field_issues_df
            ),
            "interpretation": (
                "Issues related to field IDs, duplicates, "
                "or session matching."
            ),
        },
        {
            "metric": "Question warnings",
            "value": len(
                question_warnings_df
            ),
            "interpretation": (
                "Warnings about small samples or unstable "
                "friction rankings."
            ),
        },
    ]

    return pd.DataFrame(
        summary_rows
    )


def auto_fit_and_format(
    writer: pd.ExcelWriter,
) -> None:
    """
    Apply readable formatting to every worksheet.
    """
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            maximum_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                if cell.value is None:
                    continue

                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                60,
            )


def main() -> None:
    print("Starting UX pipeline validation...")

    require_file(
        PARSED_FILE
    )

    require_file(
        METRICS_FILE
    )

    require_file(
        DIAGNOSTICS_FILE
    )

    print("Reading parsed data...")

    sessions_df = pd.read_excel(
        PARSED_FILE,
        sheet_name="Sessions",
    )

    events_df = pd.read_excel(
        PARSED_FILE,
        sheet_name="Events",
    )

    fields_df = pd.read_excel(
        PARSED_FILE,
        sheet_name="Fields",
    )

    print("Reading metrics data...")

    session_summary_df = pd.read_excel(
        METRICS_FILE,
        sheet_name="Session Summary",
    )

    question_summary_df = pd.read_excel(
        METRICS_FILE,
        sheet_name="Question Summary",
    )

    print("Reading diagnostics data...")

    friction_df = pd.read_excel(
        DIAGNOSTICS_FILE,
        sheet_name="Friction Ranking",
    )

    require_columns(
        sessions_df,
        "Sessions",
        [
            "shortIntakeId",
            "startedAt",
            "submittedAt",
        ],
    )

    require_columns(
        events_df,
        "Events",
        [
            "shortIntakeId",
            "eventOrder",
            "linkId",
            "eventTime",
        ],
    )

    require_columns(
        fields_df,
        "Fields",
        [
            "shortIntakeId",
            "fieldId",
        ],
    )

    require_columns(
        session_summary_df,
        "Session Summary",
        [
            "shortIntakeId",
            "sessionStatus",
            "submittedAt",
        ],
    )

    require_columns(
        question_summary_df,
        "Question Summary",
        [
            "questionId",
            "sessionsReached",
        ],
    )

    require_columns(
        friction_df,
        "Friction Ranking",
        [
            "questionId",
            "frictionScore",
            "sessionsReached",
        ],
    )

    prepare_datetime_column(
        sessions_df,
        "startedAt",
    )

    prepare_datetime_column(
        sessions_df,
        "submittedAt",
    )

    prepare_datetime_column(
        events_df,
        "eventTime",
    )

    prepare_datetime_column(
        session_summary_df,
        "submittedAt",
    )

    prepare_datetime_column(
        session_summary_df,
        "startedAt",
    )

    print("Checking sessions...")

    session_issues_df = (
        build_session_issues(
            sessions_df,
            session_summary_df,
            events_df,
            fields_df,
        )
    )

    print("Checking events...")

    event_issues_df = (
        build_event_issues(
            sessions_df,
            events_df,
        )
    )

    print("Checking fields...")

    field_issues_df = (
        build_field_issues(
            sessions_df,
            fields_df,
        )
    )

    print("Checking question metrics...")

    question_warnings_df = (
        build_question_warnings(
            question_summary_df,
            friction_df,
        )
    )

    validation_summary_df = (
        build_validation_summary(
            session_issues_df,
            event_issues_df,
            field_issues_df,
            question_warnings_df,
        )
    )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    print("Writing validation workbook...")

    output_sheets = {
        "Validation Summary": (
            validation_summary_df
        ),
        "Session Issues": (
            session_issues_df
        ),
        "Event Issues": (
            event_issues_df
        ),
        "Field Issues": (
            field_issues_df
        ),
        "Question Warnings": (
            question_warnings_df
        ),
    }

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        for (
            sheet_name,
            dataframe,
        ) in output_sheets.items():
            make_excel_safe(
                dataframe
            ).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        auto_fit_and_format(
            writer
        )

    total_findings = (
        len(session_issues_df)
        + len(event_issues_df)
        + len(field_issues_df)
        + len(question_warnings_df)
    )

    print()
    print("UX pipeline validation complete.")
    print(
        f"Total findings: "
        f"{total_findings}"
    )
    print(
        f"Saved to: "
        f"{OUTPUT_FILE.resolve()}"
    )


if __name__ == "__main__":
    main()